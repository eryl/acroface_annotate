import json
from pathlib import Path
from io import BytesIO
import queue
import time
import multiprocessing
#import multiprocessing.dummy as multiprocessing
import random

import numpy as np
import openpyxl

from psychopy import visual, core, event, gui #import some libraries from PsychoPy
from psychopy.hardware import keyboard

import aes
from PIL import Image

WINDOW_SIZE = (1200, 800)
IMAGE_SIZE = (400*3, 400)
N_PROCESSES = 2
MAX_QUEUE = N_PROCESSES
ENCRYPTED_DATA_DIR = Path("encrypted_data")
UNENCRYPTED_DATA_DIR = Path("data")
BACKUP_RESULTS_DIR = Path("results_backup")
BACKUP_RESULTS_DIR.mkdir(exist_ok=True, parents=True)
PAUSE_SECONDS = 60*2  # The time of the pause in seconds
PAUSE_DURATION_STEPS = 1
N_ANNOTATIONS_BEFORE_PAUSE = 500  # How many annotations to do before enforcing pause
TIME_BETWEEN_PAUSES = 60*25  # How much time to elapse between annotations
ENFORCE_ANNOTATION_PAUSE = False  # Change this to True to block the program until PAUSE_SECONDS has elapsed

TEXT_HEIGHT = 0.08

def load_file(file, bytes_password, data_is_encrypted):
    if data_is_encrypted:
        with open(file, 'rb') as encrypted_fp:
            encrypted_bytes = encrypted_fp.read()
            
            #print("Decrypting")
            t0 = time.time()
            decrypted_bytes = aes.decrypt(bytes_password, encrypted_bytes)
            dt = time.time() - t0
            #print(f"Done decrypting in {dt} seconds")
            decrypted_buffer = BytesIO(decrypted_bytes)
            image = Image.open(decrypted_buffer)
    else:
        image = Image.open(str(file))
    return file, image


def load_file_pool_worker(packed_work):
    file, bytes_password, data_is_encrypted, sem = packed_work
    result = load_file(file, bytes_password, data_is_encrypted)
    sem.acquire()
    return result


def load_file_worker(work_queue, results_queue, bytes_password, data_is_encrypted):
    try:
        file = work_queue.get_nowait()
        file, clip = load_file(file, bytes_password, data_is_encrypted)
        results_queue.put((file, clip))
    except queue.Empty:
        return


def load_images(files, bytes_password, data_is_encrypted, n_processes):
    if n_processes > 1:
        with multiprocessing.Pool(processes=N_PROCESSES) as pool, multiprocessing.Manager() as manager:
            sem = manager.Semaphore(MAX_QUEUE)
            for result in pool.imap(load_file_pool_worker, [(file, bytes_password, data_is_encrypted, sem) for file in files]):
                sem.release()
                yield result
    else:
        for file in files:
            yield load_file(file, bytes_password, data_is_encrypted)


def decrypt_work_package(work_package):
    image_path, output_directory, format, key = work_package
    with open(image_path, 'rb') as fp:
        encrypted_bytes = fp.read()
    decrypted_bytes = aes.decrypt(key, encrypted_bytes)
    with open(output_directory / image_path.name, 'wb') as fp:
        fp.write(decrypted_bytes)
    return image_path

def decrypt_file(file_path, key):
    with open(file_path, 'rb') as fp:
        encrypted_bytes = fp.read()
    decrypted_bytes = aes.decrypt(key, encrypted_bytes)
    return decrypted_bytes

def load_encrypted_image(file_path, key):
    decrypted_bytes = decrypt_file(file_path, key)
    # load audio here
    return decrypted_bytes


class EncryptedDatasetItem:
    def __init__(self, file_path, key, process_pool: multiprocessing.Pool):
        self.file_path = file_path
        self.key = key
        self.process_pool = process_pool
        self.awaited = None
        self.bytes = None
    
    def prefetch(self):
        if self.bytes is None and self.awaited is None:
            self.awaited = self.process_pool.apply_async(load_encrypted_image, (self.file_path, self.key))
    
    def get_bytes(self):
        if self.bytes is None:
            self.prefetch()
            self.bytes = self.awaited.get()
            self.awaited = None
        return self.bytes
    
    def get_image(self):
        audio_bytes = self.get_bytes()
        audio_input_buffer = BytesIO(audio_bytes)
        image = Image.open(audio_input_buffer)
        return image
    
    def clear(self):
        self.bytes = None
        self.awaited = None

def discover_data(data_directory: Path, suffix_whitelist=('.jpeg', '.png', '.enc')):
    files = [file for file in data_directory.iterdir() if file.suffix in suffix_whitelist]
    # files_listing = data_directory / 'data_file.txt'
    # if files_listing.exists():
    #     with open(files_listing) as fp:
    #         files = [data_directory / line.strip() for line in fp]
    #         return files
    # else:
    #     raise ValueError(f"Directory {data_directory} does not contain a data file listing")
    return files

class EncryptedDataset:
    def __init__(self, data_dir, password, num_processes=N_PROCESSES, prefetch_distance=N_PROCESSES, file_blacklist=None):
        self.data_dir = data_dir
        self.file_listing = discover_data(data_dir)
        if file_blacklist is not None:
            self.file_listing = [file for file in self.file_listing if file.name not in file_blacklist]
        self.key = bytes(password, encoding='utf8')
        self.num_processes = num_processes
        self.prefetch_distance = prefetch_distance
        self.pool = multiprocessing.Pool(self.num_processes)
        self.dataset_items = [EncryptedDatasetItem(file_path, self.key, self.pool) for file_path in self.file_listing]
        
    def __len__(self):
        return len(self.file_listing)
    
    def __getitem__(self, index):
        for i in range(index):
            self.dataset_items[i].clear()
            
        self.prefetch(index)    
        t0 = time.time()
        bytes = self.dataset_items[index].get_bytes()  # will block if bytes has not been read
        print(f"Bytes took {time.time()-t0}s to read")
        return bytes
    
    def get_image(self, index):
        image_bytes = self[index]
        return image_from_bytes(image_bytes)

    def get_file_name(self, index):
        return self.dataset_items[index].file_path
    
    def cache(self, index):
        """Initiate a prefectch of the given index. The encryption latency will be hidden since it's likely done ahead of time"""
        pass
    
    def evict(self, index):
        """If the given index has been cached, destroy the cached object"""
        pass
    
    def shutdown_pool(self):
        self.pool.close()
        self.pool.join()
        
    def __del__(self):
        self.shutdown_pool()
    
    def shuffle(self, seed=1729):
        random.seed(seed)
        random.shuffle(self.dataset_items)
        self.prefetch()
            
    def prefetch(self, index=0):
        for i in range(index, index + self.prefetch_distance + 1):
            if i < len(self):
                self.dataset_items[i].prefetch()
        

def image_from_bytes(image_bytes):
    image_input_buffer = BytesIO(image_bytes)
    sound = Image.open(image_input_buffer)
    return sound


def test_decrypt(file_directory: Path, password):
    key = bytes(password, encoding='utf8')
    file_listing = discover_data(file_directory)
    try:
        decrypt_file(file_listing[0], key)
        return True
    except AssertionError:
        return False
    # except InvalidToken:
    #     return False


def main():
    data_files = discover_data(ENCRYPTED_DATA_DIR)
    ## Ask for decryption key
    data_is_encrypted = True
    data_read_success = False
    while not data_read_success:
        dlg = gui.Dlg(title="Alternativ för kryptering")
        # Add each field manually
        dlg.addField('Lösenord*', tip='Lösenord som används för att avkryptera data. Lämna tomt om data är okrypterat.')

        thisInfo = dlg.show()
        if thisInfo is not None:
            password, = thisInfo
            bytes_password = bytes(password, encoding='utf8')
            
            if test_decrypt(ENCRYPTED_DATA_DIR, password):
                data_read_success = True
            else:
                gui.warnDlg(title="Felaktigt lösenord", prompt=f"Lösenordet som angavs kunde inte avkryptera filen {data_files[0]} , vänligen försök igen.")
        else:
            #mywin.close()
            core.quit()
            
        
    partial_results_file = Path("partial_results.xlsx")

    fieldnames = ["file", 'annotation', 'n_changes', 'decision_time']

    annotations = dict()
    if partial_results_file.exists():
        wb = openpyxl.load_workbook(filename = partial_results_file)
        ws = wb.active
        rows = list(ws.iter_rows())
        # We assume the first row is the header
        for file, annotation, n_changes, decision_time in rows[1:]:
            annotations[file.value] = {'file': file.value, 'annotation': annotation.value, 'n_changes': n_changes.value, 'decision_time': decision_time.value}
        wb.close()

    if annotations:
        dlg_prev = gui.Dlg(title="Fortsätt med existerande annoteringar")
        # Add each field manually
        dlg_prev.addText(f"Det finns {len(annotations)} tidigare annoteringar lagrade, vill du fortsätta med dessa eller börja annotera på nytt?")
        dlg_prev.addField("Behåll annoteringar*", initial=True, required=True)
        #dlg_prev.addText(f"Ange \"OK\" för att fortsätta med tidigare annoteringar eller \"Cancel\" för att börja på nytt och radera tidigare annoteringar")
        prev_info = dlg_prev.show()
        if prev_info is not None:
            keep_annotations, = prev_info
            if not keep_annotations:
                dlg_prev = gui.Dlg(title="Bekräfta radering av existerande annoteringar")
                # Add each field manually
                dlg_prev.addText(f"Om du fortsätter kommer {len(annotations)} existerande annoteringar att raderas.")
                dlg_prev.addText(f"Ange \"OK\" för radera tidigare annoteringar eller \"Cancel\" för att fortsätta utan att radera")
                prev_info = dlg_prev.show()
                if prev_info is not None:
                    partial_results_file.unlink()
                    annotations = dict()
        else:
            core.quit() 
            
    dataset = EncryptedDataset(ENCRYPTED_DATA_DIR, password, file_blacklist=set(annotations.keys()))
    dataset.shuffle()

    #create a window
    mywin = visual.Window(WINDOW_SIZE, monitor="testMonitor", units="deg")

    annotation_stim = visual.TextStim(mywin, 
                                    text="",
                                    pos=(0.0, 0.8),
                                    units="norm",
                                    height=TEXT_HEIGHT,
                                    wrapWidth=1.6,)
                                    
    instruction_stim = visual.TextStim(mywin, 
                                    text="Tryck på 'A' för akromegali eller 'K' för kontroll\n"
                                         "Om du känner igen personen - tryck 'X'\n\n"
                                         "Avbryt genom att trycka på Escape eller 'Q' ",
                                    pos=(0.0, -0.7),
                                    units="norm",
                                    height=TEXT_HEIGHT,
                                    wrapWidth=1.6,)
    #create a keyboard component
    kb = keyboard.Keyboard()

    # Flag to exit the experiment early, but still save intermediate results
    quit_experiment = False

    ## This set's up the annotation commands
    annotation_commands = {'a': {'name': 'Akromegali', 'color': (.5, 0, 0)}, 
                        'k': {'name': 'Kontroll', 'color': (0, 0, .5)}, 
                        'x': {'name': 'Känd', 'color': (0, .5, .5)}}

    n_images = len(dataset)
    
    annotations_since_last_pause = 0
    last_pause_time = time.time()
    
    for i in range(n_images):
        elapsed_since_pause = time.time() - last_pause_time
        if (annotations_since_last_pause > N_ANNOTATIONS_BEFORE_PAUSE) or (elapsed_since_pause > TIME_BETWEEN_PAUSES):
            if ENFORCE_ANNOTATION_PAUSE:
                mywin.color = (0, 0, 0)
                mywin.flip()
                pause_duration = PAUSE_SECONDS
                sleep_time = 1
                
                if pause_duration < 60:
                    pause_text = f"Ta en paus i {pause_duration}s"
                else:
                    minutes = pause_duration // 60
                    seconds = pause_duration % 60
                    pause_text = f"Ta en paus i {minutes}m:{seconds}s"
                    
                pause_stim = visual.TextStim(mywin, 
                                        text=pause_text,
                                        pos=(0.0, 0.0),
                                        units="norm",
                                        height=TEXT_HEIGHT,
                                        wrapWidth=1.6,)
                
                while pause_duration > 0:
                    key_presses = kb.getKeys(clear=True, waitRelease=False)
                    for key_press in key_presses:
                        key_name = key_press.name
                        if key_name in ['q', 'escape']:
                            annotation_done = True
                            quit_experiment = True    
                        else:
                            continue
                    if quit_experiment:
                        break
                    
                    pause_stim.draw()
                    mywin.flip()
                    #time.sleep(sleep_time)
                    core.wait(sleep_time)
                    pause_duration -= PAUSE_DURATION_STEPS
                    if pause_duration < 60:
                        pause_text = f"Ta en paus i {pause_duration}s"
                    else:
                        minutes = pause_duration // 60
                        seconds = pause_duration % 60
                        pause_text = f"Ta en paus i {minutes}m:{seconds}s"
                    pause_stim.text = pause_text

                if quit_experiment:
                    break
                
                gui.infoDlg(prompt="Redo att starta igen")
            else:
                
                pause_text = (f"Vi rekommenderar att du tar en fem minuters paus.\n"
                               "Tryck Enter för att fortsätta.\n"
                               "Tryck på 'Q' eller Escape avslutar.")
                pause_stim = visual.TextStim(mywin, 
                                        text=pause_text,
                                        pos=(0.0, 0.0),
                                        units="norm",
                                        height=TEXT_HEIGHT,
                                        wrapWidth=1.6,)
                do_pause = True
                while do_pause:
                    mywin.color = (-1., -1., -1.)
                    pause_stim.draw()
                    mywin.flip()
                    
                    key_presses = kb.getKeys(clear=True, waitRelease=False)
                    for key_press in key_presses:
                        key_name = key_press.name
                        if key_name in ['q', 'escape']:
                            annotation_done = True
                            quit_experiment = True
                            do_pause = False
                        elif key_name == 'return':
                            do_pause = False
                        else:
                            continue
                    time.sleep(0.2)
                    
                if quit_experiment:
                    break
                    
            annotations_since_last_pause = 0
            last_pause_time = time.time()
        
        
        pil_image = dataset.get_image(i)
        file = dataset.get_file_name(i)
        image_arr = np.array(pil_image, order="C")  # convert to numpy array with shape width, height, channels
        image_arr = (image_arr.astype(float) / 255.0)  # convert to float in 0--1 range, assuming image is 8-bit uint.

        image_stim = visual.ImageStim(mywin, 
                                    image_arr[::-1],
                                    units="pix",
                                    pos=(0, 0),
                                    size=IMAGE_SIZE,  # here's a gotcha: need to pass the size (x, y) explicitly.
                                    colorSpace="rgb1")  # img_as_float converts to 0:1 range, whereas PsychoPy defaults to -1:1.
        
        annotation_stim.text = ""
        mywin.color = (.2, .2, .2)
        mywin.flip()
        #draw the stimuli and update the window
        annotation = 'none'
        n_changes = 0
        trialClock = core.Clock()
        
            
        annotation_done = False
        while not annotation_done:
            annotation_stim.draw()
            instruction_stim.draw()
            image_stim.draw()
            mywin.flip()

            key_presses = kb.getKeys(clear=True, waitRelease=False)
            for key_press in key_presses:
                key_name = key_press.name
                if key_name in annotation_commands:
                    if annotation != key_name:
                        n_changes += 1
                    annotation = key_name
                    annotation_data = annotation_commands[key_name]
                    annotation_stim.text = f"Din notering: {annotation_data['name']}\nTryck 'enter' för att bekräfta."
                    mywin.color = annotation_data['color']
                    mywin.flip()

                elif key_press == 'return' and n_changes > 0:
                    annotation_done = True
                    break
                
                elif key_name in ['q', 'escape']:
                    annotation_done = True
                    quit_experiment = True
                    
                else:
                    continue
                
        if quit_experiment:
            break

        event.clearEvents()

        decision_time = trialClock.getTime()
        annotation_result = {'file': file.name, 'annotation': annotation, 'n_changes': n_changes, 'decision_time': decision_time}
        annotations[file.name] = annotation_result
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(fieldnames)
        for file_name, record in annotations.items():
            row = [record[fieldname] for fieldname in fieldnames]
            ws.append(row)
        wb.save(partial_results_file)
        backup_results_file = BACKUP_RESULTS_DIR / file.with_suffix('.json').name
        with open(backup_results_file ,'w') as backup_results_fp:
            json.dump(annotation_result, backup_results_fp)
            
        annotations_since_last_pause += 1
        # with open(partial_results_file, 'a', newline='') as partial_results_fp:
        #     csv_writer = csv.DictWriter(partial_results_fp, fieldnames=fieldnames)
        #     csv_writer.writerow({'file': file.name, 'annotation': annotation, 'n_changes': n_changes, 'decision_time': decision_time})
        
    if set([f.name for f in data_files]).issubset(annotations.keys()):
        file_save_results = gui.fileSaveDlg(prompt="Välj fil att exportera resultat till",initFileName="results.xlsx", allowed='Excel files (*.xlsx)|*.xlsx')
        if file_save_results is None:
            gui.infoDlg(title="Resultat sparades inte", prompt=f"Resultatet sparades inte, kör programmet igen (dina annoteringar har sparats) och välj fil att spara till.")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(fieldnames)
            for file_name, record in annotations.items():
                row = [record[fieldname] for fieldname in fieldnames]
                ws.append(row)
            wb.save(file_save_results)
            #trials.saveAsExcel(file_save_results, dataOut=('all_raw',), fileCollisionMethod='overwrite')
    else:
        gui.infoDlg(title="Ofullständigt resultat", prompt=f"Alla filer har inte annoterats. Du kan fortsätta annotera de kvarvarande genom att köra programmet igen.")

    mywin.close()
    core.quit()


if __name__ == '__main__':
    main()